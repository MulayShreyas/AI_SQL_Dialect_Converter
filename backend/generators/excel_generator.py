from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from typing import List, Dict
from datetime import datetime
import io

class ExcelGenerator:
    """
    Generator for creating Excel files with SQL conversion results.
    Uses openpyxl for xlsx file generation.
    """
    
    def __init__(self):
        # Define styles
        self.header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
        self.header_font = Font(color='FFFFFF', bold=True, size=11)
        self.success_fill = PatternFill(start_color='d1fae5', end_color='d1fae5', fill_type='solid')
        self.error_fill = PatternFill(start_color='fee2e2', end_color='fee2e2', fill_type='solid')
        self.border = Border(
            left=Side(style='thin', color='e0e0e0'),
            right=Side(style='thin', color='e0e0e0'),
            top=Side(style='thin', color='e0e0e0'),
            bottom=Side(style='thin', color='e0e0e0')
        )
        self.wrap_alignment = Alignment(wrap_text=True, vertical='top')
    
    def generate(
        self, 
        results: List[Dict], 
        source_dialect: str, 
        target_dialect: str,
        filename: str = None
    ) -> io.BytesIO:
        """
        Generate an Excel file with conversion results.
        
        Args:
            results: List of conversion results
            source_dialect: Source SQL dialect
            target_dialect: Target SQL dialect
            filename: Optional filename (for metadata)
            
        Returns:
            BytesIO buffer containing the Excel file
        """
        wb = Workbook()
        
        # Create summary sheet
        self._create_summary_sheet(wb, results, source_dialect, target_dialect)
        
        # Create detailed results sheet
        self._create_results_sheet(wb, results, source_dialect, target_dialect)
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    def _create_summary_sheet(
        self, 
        wb: Workbook, 
        results: List[Dict],
        source_dialect: str,
        target_dialect: str
    ):
        """Create the summary sheet."""
        ws = wb.active
        ws.title = "Summary"
        
        # Title
        ws['A1'] = "SQL Dialect Conversion Report"
        ws['A1'].font = Font(size=18, bold=True, color='1a1a2e')
        ws.merge_cells('A1:D1')
        
        # Metadata
        ws['A3'] = "Source Dialect:"
        ws['B3'] = source_dialect
        ws['A4'] = "Target Dialect:"
        ws['B4'] = target_dialect
        ws['A5'] = "Generated:"
        ws['B5'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Style metadata
        for row in range(3, 6):
            ws[f'A{row}'].font = Font(bold=True)
        
        # Statistics
        ws['A7'] = "Statistics"
        ws['A7'].font = Font(size=14, bold=True, color='667eea')
        ws.merge_cells('A7:B7')
        
        success_count = sum(1 for r in results if r['status'] == 'success')
        error_count = len(results) - success_count
        success_rate = (success_count / len(results) * 100) if results else 0
        
        stats = [
            ("Total Statements", len(results)),
            ("Successfully Converted", success_count),
            ("Errors", error_count),
            ("Success Rate", f"{success_rate:.1f}%")
        ]
        
        for i, (label, value) in enumerate(stats, 9):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
            
            # Color code success/error rows
            if "Success" in label and isinstance(value, int):
                ws[f'B{i}'].fill = self.success_fill
            elif "Error" in label:
                ws[f'B{i}'].fill = self.error_fill
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
    
    def _create_results_sheet(
        self, 
        wb: Workbook, 
        results: List[Dict],
        source_dialect: str,
        target_dialect: str
    ):
        """Create the detailed results sheet."""
        ws = wb.create_sheet("Conversion Results")
        
        # Headers
        headers = ["#", "Status", f"Original SQL ({source_dialect})", 
                   f"Converted SQL ({target_dialect})", "Notes"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        for row_idx, result in enumerate(results, 2):
            # Row number
            ws.cell(row=row_idx, column=1, value=row_idx - 1)
            
            # Status
            status_cell = ws.cell(row=row_idx, column=2)
            if result['status'] == 'success':
                status_cell.value = "✓ Success"
                status_cell.fill = self.success_fill
            else:
                status_cell.value = "✗ Error"
                status_cell.fill = self.error_fill
            
            # Original SQL
            ws.cell(row=row_idx, column=3, value=result['original'])
            
            # Converted SQL
            ws.cell(row=row_idx, column=4, value=result.get('converted', ''))
            
            # Notes
            ws.cell(row=row_idx, column=5, value=result.get('notes', ''))
            
            # Apply styling to all cells in row
            for col in range(1, 6):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = self.border
                cell.alignment = self.wrap_alignment
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 40
        
        # Freeze header row
        ws.freeze_panes = 'A2'
