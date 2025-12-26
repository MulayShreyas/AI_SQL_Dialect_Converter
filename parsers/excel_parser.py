"""
Excel Parser - Extracts SQL statements from Excel files
"""

import pandas as pd
from openpyxl import load_workbook
from typing import List
import io
import re


class ExcelParser:
    """
    Parser for extracting SQL statements from Excel files.
    Handles both .xlsx and .xls formats.
    """
    
    def __init__(self):
        self.sql_keywords = [
            'SELECT', 'INSERT', 'UPDATE', 'DELETE', 'CREATE', 'ALTER', 
            'DROP', 'WHERE', 'FROM', 'JOIN', 'UNION', 'WITH'
        ]
    
    def parse(self, file) -> List[str]:
        """
        Extract SQL statements from an Excel file.
        
        Args:
            file: File-like object or path to Excel file
            
        Returns:
            List of SQL statements found in the Excel file
        """
        all_text = self._extract_all_cells(file)
        return self._extract_sql_statements(all_text)
    
    def _extract_all_cells(self, file) -> List[str]:
        """Extract text from all cells in all sheets."""
        cell_values = []
        
        try:
            # Handle file-like object
            if hasattr(file, 'read'):
                file_bytes = file.read()
                file.seek(0)  # Reset for potential reuse
                excel_file = io.BytesIO(file_bytes)
            else:
                excel_file = file
            
            # Read all sheets
            xl = pd.ExcelFile(excel_file)
            
            for sheet_name in xl.sheet_names:
                df = pd.read_excel(xl, sheet_name=sheet_name, header=None)
                
                # Extract each cell value
                for col in df.columns:
                    for value in df[col].dropna():
                        str_value = str(value).strip()
                        if str_value:
                            cell_values.append(str_value)
            
        except Exception as e:
            print(f"Error reading Excel file: {e}")
            # Try alternative method with openpyxl
            cell_values = self._extract_with_openpyxl(file)
        
        return cell_values
    
    def _extract_with_openpyxl(self, file) -> List[str]:
        """Fallback extraction using openpyxl directly."""
        cell_values = []
        
        try:
            if hasattr(file, 'read'):
                file.seek(0)
                wb = load_workbook(filename=io.BytesIO(file.read()), data_only=True)
            else:
                wb = load_workbook(filename=file, data_only=True)
            
            for sheet in wb.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value:
                            str_value = str(cell.value).strip()
                            if str_value:
                                cell_values.append(str_value)
            
            wb.close()
        except Exception as e:
            print(f"Error with openpyxl: {e}")
        
        return cell_values
    
    def _extract_sql_statements(self, cell_values: List[str]) -> List[str]:
        """
        Extract SQL statements from cell values.
        
        Handles both:
        - Complete SQL statements in single cells
        - SQL spread across multiple cells
        """
        statements = []
        combined_text = ""
        
        for value in cell_values:
            # Check if this cell contains SQL
            if self._contains_sql(value):
                # Clean and add the statement
                cleaned = self._clean_sql_statement(value)
                if cleaned:
                    statements.append(cleaned)
            
            # Also combine for multi-cell SQL
            combined_text += " " + value
        
        # Look for SQL in combined text if individual cells didn't have complete statements
        if len(statements) == 0:
            combined_statements = self._extract_from_combined(combined_text)
            statements.extend(combined_statements)
        
        return self._deduplicate_statements(statements)
    
    def _contains_sql(self, text: str) -> bool:
        """Check if text contains SQL content."""
        upper_text = text.upper()
        
        # Must contain SQL keywords
        keyword_count = sum(1 for kw in self.sql_keywords if kw in upper_text)
        
        # Need at least 2 keywords for a likely SQL statement
        return keyword_count >= 2
    
    def _clean_sql_statement(self, statement: str) -> str:
        """Clean and format a SQL statement."""
        # Remove excessive whitespace
        statement = re.sub(r'\s+', ' ', statement.strip())
        
        # Ensure ends with semicolon
        if statement and not statement.endswith(';'):
            statement += ';'
        
        return statement
    
    def _extract_from_combined(self, text: str) -> List[str]:
        """Extract SQL from combined cell text."""
        statements = []
        
        # Clean the text
        text = re.sub(r'\s+', ' ', text.strip())
        
        # Pattern for SQL statements
        sql_pattern = r'(?i)((?:SELECT|INSERT|UPDATE|DELETE|CREATE|ALTER|DROP|WITH)\b[^;]*(?:;|$))'
        
        matches = re.findall(sql_pattern, text, re.DOTALL | re.IGNORECASE)
        
        for match in matches:
            cleaned = self._clean_sql_statement(match)
            if cleaned and len(cleaned) > 15:  # Minimum viable SQL length
                statements.append(cleaned)
        
        return statements
    
    def _deduplicate_statements(self, statements: List[str]) -> List[str]:
        """Remove duplicate statements while preserving order."""
        seen = set()
        unique = []
        
        for stmt in statements:
            # Normalize for comparison
            normalized = stmt.upper().strip()
            if normalized not in seen:
                seen.add(normalized)
                unique.append(stmt)
        
        return unique
